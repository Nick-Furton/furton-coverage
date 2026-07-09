#!/usr/bin/env python
"""Build a per-ticker Excel model from SEC EDGAR data (Session 3).

Pulls a ticker's XBRL companyfacts through ``scripts/edgar.py`` (the only module that
touches the network) and lays out a multi-tab workbook:

* **Cover**        -- ticker/company metadata, data sources, color legend, disclaimer.
* **Income Statement** -- trailing 8 quarters of headline financials (revenue, gross
  profit, operating income, net income, diluted EPS) from XBRL companyfacts, with
  margin/growth rows as live Excel formulas (never pre-computed in Python).
* **KPIs**         -- the company-specific KPIs listed in config/universe.json
  (segment revenue, ARR, procedure growth, etc.). companyfacts does NOT carry these
  (Session 2's structural finding -- see PLAN §3 / universe.json _schema.companyfacts_caveat),
  so this tab is TBD-flagged (yellow) unless real, cited figures are supplied via
  ``--kpi-overrides`` (a small JSON file of press-release-extracted values with their
  source URL) -- never fabricated.
* **Our Estimates** -- forward estimate for the next reported period. Assumption cells
  are the only hardcoded inputs (blue); the estimate itself is a live formula
  (black) built off a green cross-sheet reference to the last actual on the Income
  Statement tab -- editable, re-runnable, not a one-shot snapshot.

Because this model is built entirely from public SEC data (XBRL + press-release
extraction we did ourselves), it is 100% ours to publish (PLAN §1/§8) -- no vendor
licensing question.

Color convention (xlsx skill / IB standard): blue = hardcoded input, black = formula,
green = cross-sheet reference, yellow fill = needs attention / TBD.

Windows note: no LibreOffice on this machine, so ``scripts/recalc.py``-style formula
recalculation isn't available here -- formulas are written as live Excel formulas and
will compute correctly the moment the workbook is opened in Excel (or any spreadsheet
app). ``_selfcheck_income_statement`` checks the underlying DATA is finite and that the
margin formulas' denominators are div-by-zero-guarded (they use ``IF(rev=0,"-",...)``
in the formula text itself); it does not re-execute the Excel formulas outside Excel,
so a wrong cell REFERENCE (as opposed to a bad denominator) would not be caught here --
open the workbook and spot-check 2-3 cells before trusting a new ticker's model.

CLI::

    py scripts/build_model.py TICKER [--kpi-overrides path.json] [--periods N] [--out path]
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import edgar

REPO_ROOT = Path(__file__).resolve().parent.parent
UNIVERSE_PATH = REPO_ROOT / "config" / "universe.json"
MODELS_DIR = REPO_ROOT / "models"

# --------------------------------------------------------------------------- #
# Style constants
# --------------------------------------------------------------------------- #

FONT_NAME = "Calibri"
BLUE = "0000FF"     # hardcoded input
BLACK = "000000"    # formula
GREEN = "008000"    # cross-sheet reference
GREY = "808080"      # source footnote
YELLOW_FILL = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
HEADER_FILL = PatternFill("solid", start_color="1F2937", end_color="1F2937")

CUR_FMT = "$#,##0;($#,##0);-"
EPS_FMT = "$0.00;($0.00);-"
PCT_FMT = "0.0%;(0.0%);-"

F_INPUT = Font(name=FONT_NAME, color=BLUE)
F_FORMULA = Font(name=FONT_NAME, color=BLACK)
F_XREF = Font(name=FONT_NAME, color=GREEN)
F_SOURCE = Font(name=FONT_NAME, color=GREY, italic=True, size=8)
F_HEADER = Font(name=FONT_NAME, color="FFFFFF", bold=True)
F_LABEL = Font(name=FONT_NAME, bold=True)
F_TITLE = Font(name=FONT_NAME, bold=True, size=14)


def _sheet(wb: Workbook, name: str) -> Worksheet:
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    return ws


def _header_row(ws: Worksheet, row: int, labels: list[str], start_col: int = 1) -> None:
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=label)
        c.font = F_HEADER
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")


# --------------------------------------------------------------------------- #
# Data extraction (all reads go through edgar.py; PURE beyond that point)
# --------------------------------------------------------------------------- #


def load_universe_entry(ticker: str) -> dict:
    """Best-effort read of config/universe.json for metadata. Tolerant of absence --
    build_model.py must work for any resolvable ticker, not only the frozen 10."""
    if not UNIVERSE_PATH.exists():
        return {}
    try:
        data = json.loads(UNIVERSE_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    for entry in data.get("universe", []):
        if entry.get("ticker", "").upper() == ticker.upper():
            return entry
    return {}


def _parse_date(s: str | None):
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def _duration_days(fact: "edgar.Fact") -> int | None:
    start, end = _parse_date(fact.start), _parse_date(fact.end)
    if start is None or end is None:
        return None
    return (end - start).days


def _dedup_by_end(series: list) -> tuple[dict, dict]:
    """Dedupe a Fact series by period-end. Returns (by_end, conflicts).

    XBRL companyfacts repeats every period-end as the prior-year comparative column in
    the FOLLOWING year's 10-Q -- e.g. Q1 FY26 (end 2025-04-30) shows up again inside the
    Q1 FY27 10-Q's comparative table, but re-tagged with THAT filing's own fy/fp context
    (fy=2027, fp=Q1) rather than its own. Picking the latest-filed duplicate (as if it
    were a restatement) therefore silently mislabels every period. Fix: prefer the
    EARLIEST-filed fact for a given end -- that is the quarter's own originating 10-Q,
    correctly labeled. If two duplicates for the same end genuinely disagree on value
    (a real restatement, or -- as with AMZN's 2022 stock split -- a share-count-driven
    EPS restatement), record it in ``conflicts`` for the caller to warn on, but do not
    silently pick either -- we always use the earliest-filed (as-originally-reported)
    figure for consistency."""
    by_end: dict[str, "edgar.Fact"] = {}
    conflicts: dict[str, tuple] = {}
    for f in series:
        prior = by_end.get(f.end)
        if prior is None:
            by_end[f.end] = f
            continue
        earliest, latest = (f, prior) if (f.filed or "") < (prior.filed or "") else (prior, f)
        if earliest.val != latest.val:
            conflicts[f.end] = (earliest.filed, earliest.val, latest.filed, latest.val)
        by_end[f.end] = earliest
    return by_end, conflicts


def _quarterly_by_end(facts_json: dict, tags: list[str]) -> tuple[str, dict, dict]:
    """Every duration-matched (~75-100 day) quarterly fact for a metric, deduped by end,
    with no truncation -- callers align to whatever window they need."""
    tag, series = edgar.named_metric(facts_json, tags)
    if not series:
        return tag, {}, {}
    quarterly = [f for f in series if (d := _duration_days(f)) is not None and 75 <= d <= 100]
    by_end, conflicts = _dedup_by_end(quarterly)
    return tag, by_end, conflicts


# --------------------------------------------------------------------------- #
# Cover tab
# --------------------------------------------------------------------------- #


def build_cover(wb: Workbook, ticker: str, cik_info: "edgar.CikInfo", entry: dict, generated: str) -> None:
    ws = _sheet(wb, "Cover")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    ws["A1"] = f"{ticker} -- Furton Coverage Model"
    ws["A1"].font = F_TITLE
    rows = [
        ("Company", entry.get("company") or cik_info.title),
        ("Ticker", ticker),
        ("CIK", cik_info.cik10),
        ("Sector", entry.get("sector", "")),
        ("Fiscal year end", entry.get("fiscal_year_end", "")),
        ("Guidance basis", entry.get("basis", "")),
        ("Generated", generated),
        ("Data source", "SEC EDGAR XBRL companyfacts + earnings 8-K (EX-99.1) press releases"),
        ("Companyfacts URL", edgar.COMPANYFACTS_URL.format(cik10=cik_info.cik10)),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = F_LABEL
        ws.cell(row=r, column=2, value=value).font = F_FORMULA
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Color legend").font = F_LABEL
    r += 1
    legend = [
        ("Blue text", "Hardcoded input (XBRL actual, or an assumption we set)", F_INPUT),
        ("Black text", "Formula / calculation", F_FORMULA),
        ("Green text", "Cross-sheet reference (pulls from another tab)", F_XREF),
        ("Yellow fill", "TBD -- not in companyfacts; needs press-release extraction", None),
    ]
    for label, desc, font in legend:
        c = ws.cell(row=r, column=1, value=label)
        if font:
            c.font = font
        else:
            c.fill = YELLOW_FILL
        ws.cell(row=r, column=2, value=desc).font = F_SOURCE
        r += 1
    r += 1
    ws.cell(
        row=r, column=1,
        value="Educational research generated from public SEC EDGAR filings. Not investment advice.",
    ).font = F_SOURCE


# --------------------------------------------------------------------------- #
# Income Statement tab
# --------------------------------------------------------------------------- #

METRIC_ROWS = [
    ("Revenue", edgar.REVENUE_TAGS),
    ("Gross profit", edgar.GROSS_PROFIT_TAGS),
    ("Operating income", edgar.OPERATING_INCOME_TAGS),
    ("Net income", edgar.NET_INCOME_TAGS),
    ("Diluted EPS", edgar.EPS_DILUTED_TAGS),
]


_MONTH_NUM = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}


def _fye_month(fiscal_year_end: str | None) -> int | None:
    """Parse a single clean month name (e.g. 'January') to its number. Ambiguous/compound
    strings (e.g. MU's 'August/September') return None -- callers fall back to raw dates
    rather than guessing which month anchors the fiscal year."""
    if not fiscal_year_end:
        return None
    key = fiscal_year_end.strip().lower()
    return _MONTH_NUM.get(key)


def _period_label(end_iso: str, fye_month: int | None) -> str:
    """Compute 'Q{n} FY{yy}' purely from the period-end date + the company's fiscal-year-end
    month -- NOT from the fact's self-reported fy/fp. Those XBRL fields are unreliable:
    companyfacts re-tags a period's PRIOR-YEAR comparative reappearance in a later filing
    with that later filing's own fy context, and some filers' dei:DocumentFiscalYearFocus
    is inconsistent year over year even for the ORIGINAL filing of a period (verified via
    CRWD: raw fy on the originating 10-Q for both 2024-04-30 and 2025-04-30 -- a full fiscal
    year apart -- was identically 'FY2025', while the company's own press release calls the
    analogous 2026-04-30 quarter 'first quarter fiscal year 2027', confirming the raw field
    drifts and a purely date-derived label is the trustworthy one)."""
    end = _parse_date(end_iso)
    if end is None or fye_month is None:
        return end_iso
    offset = (12 - fye_month) % 12
    shifted_month = ((end.month - 1 + offset) % 12) + 1
    quarter = (shifted_month - 1) // 3 + 1
    fy = end.year + 1 if end.month > fye_month else end.year
    return f"Q{quarter} FY{str(fy)[-2:]}"


def build_income_statement(
    wb: Workbook,
    facts_json: dict,
    ticker: str,
    n_periods: int,
    fiscal_year_end: str | None,
    cik10: str,
) -> dict:
    """Returns {metric_label: {period_end: Fact}} for reuse by the Estimates tab."""
    ws = _sheet(wb, "Income Statement")
    ws.freeze_panes = "B4"
    fye_month = _fye_month(fiscal_year_end)

    # Anchor every row to REVENUE's period-end set (the one tag virtually every filer
    # keeps current). A metric whose tag went stale/renamed (e.g. DE stopped tagging a
    # standalone GrossProfit/OperatingIncomeLoss in recent 10-Qs -- Session 2's
    # segment-data finding, PLAN §3) must NOT reappear under a much-older, misaligned
    # column; it is better to show a blank cell than 2019 data mislabeled as 2026.
    anchor_tag, anchor_by_end, anchor_conflicts = _quarterly_by_end(facts_json, METRIC_ROWS[0][1])
    if not anchor_by_end:
        raise edgar.EdgarError(
            f"{ticker}: no quarterly-duration Revenue fact found under any of "
            f"{METRIC_ROWS[0][1]} -- cannot anchor the Income Statement tab. This is a "
            "different, more severe data-quality issue than DE/CRWD's stale-tag case "
            "(where only a downstream metric went stale); fix loud rather than emit a "
            "blank sheet."
        )
    ends_sorted = sorted(anchor_by_end)[-n_periods:]
    for end in ends_sorted:
        if end in anchor_conflicts:
            ef, ev, lf, lv = anchor_conflicts[end]
            print(
                f"  [build_model] warning: {anchor_tag} at {end} disagrees between filings "
                f"({ef}={ev} vs {lf}={lv}) -- using earliest-filed",
                file=sys.stderr,
            )

    by_metric_by_end: dict[str, dict] = {METRIC_ROWS[0][0]: {e: anchor_by_end[e] for e in ends_sorted}}
    for label, tags in METRIC_ROWS[1:]:
        tag, by_end, conflicts = _quarterly_by_end(facts_json, tags)
        by_metric_by_end[label] = {e: by_end[e] for e in ends_sorted if e in by_end}
        covered = sum(1 for e in ends_sorted if e in by_end)
        if covered == 0:
            last_avail = max(by_end) if by_end else None
            print(
                f"  [build_model] warning: {label} ({tag}) has no data in the displayed "
                f"{len(ends_sorted)}-quarter window (last tagged period: {last_avail or 'none'}) "
                "-- likely a dropped/renamed XBRL concept; row left blank, see the 8-K for this figure",
                file=sys.stderr,
            )
        elif covered < len(ends_sorted):
            print(
                f"  [build_model] warning: {label} ({tag}) covers only {covered}/{len(ends_sorted)} "
                "of the displayed quarters -- some cells will be blank",
                file=sys.stderr,
            )
        for end in ends_sorted:
            if end in by_end and end in conflicts:
                ef, ev, lf, lv = conflicts[end]
                print(
                    f"  [build_model] warning: {tag} at {end} disagrees between filings "
                    f"({ef}={ev} vs {lf}={lv}) -- using earliest-filed",
                    file=sys.stderr,
                )

    ws["A1"] = f"{ticker} -- Income Statement (trailing {len(ends_sorted)} quarters, XBRL companyfacts)"
    ws["A1"].font = F_TITLE
    ws.column_dimensions["A"].width = 22
    header_row = 3
    ws.cell(row=header_row, column=1, value="Metric ($M except EPS)").font = F_HEADER
    ws.cell(row=header_row, column=1).fill = HEADER_FILL
    period_labels = [_period_label(end, fye_month) for end in ends_sorted]
    _header_row(ws, header_row, period_labels, start_col=2)

    row_index: dict[str, int] = {}
    r = header_row + 1
    for label, _tags in METRIC_ROWS:
        ws.cell(row=r, column=1, value=label).font = F_LABEL
        row_index[label] = r
        for i, end in enumerate(ends_sorted):
            col = 2 + i
            f = by_metric_by_end[label].get(end)
            cell = ws.cell(row=r, column=col)
            if f is None:
                cell.value = None
                continue
            if f.val is None:
                raise ValueError(
                    f"{ticker}: {label} fact at {end} (accn {f.accn}) has a null 'val' -- "
                    "malformed XBRL units row; fail loud rather than crash on the arithmetic "
                    "below or silently write a blank/wrong cell"
                )
            is_eps = label == "Diluted EPS"
            # revenue/profit tags are USD; scale to $M for readability, EPS stays as-is.
            cell.value = round(f.val, 2) if is_eps else round(f.val / 1_000_000, 2)
            cell.font = F_INPUT
            cell.number_format = EPS_FMT if is_eps else CUR_FMT
        r += 1

    # Margin rows -- live formulas, never pre-computed in Python (xlsx skill rule).
    gm_row, om_row, nm_row = r, r + 1, r + 2
    ws.cell(row=gm_row, column=1, value="Gross margin %").font = F_LABEL
    ws.cell(row=om_row, column=1, value="Operating margin %").font = F_LABEL
    ws.cell(row=nm_row, column=1, value="Net margin %").font = F_LABEL
    rev_r, gp_r, oi_r, ni_r = (row_index["Revenue"], row_index["Gross profit"],
                               row_index["Operating income"], row_index["Net income"])
    for i in range(len(ends_sorted)):
        col = get_column_letter(2 + i)
        for target_row, num_row in ((gm_row, gp_r), (om_row, oi_r), (nm_row, ni_r)):
            cell = ws.cell(row=target_row, column=2 + i)
            cell.value = f"=IF({col}{rev_r}=0,\"-\",{col}{num_row}/{col}{rev_r})"
            cell.font = F_FORMULA
            cell.number_format = PCT_FMT
    r = nm_row + 1

    # YoY revenue growth -- another live formula, offset 4 columns back (4 quarters).
    yoy_row = r + 1
    ws.cell(row=yoy_row, column=1, value="Revenue YoY growth %").font = F_LABEL
    for i in range(len(ends_sorted)):
        col_letter = get_column_letter(2 + i)
        cell = ws.cell(row=yoy_row, column=2 + i)
        if i >= 4:
            prior_letter = get_column_letter(2 + i - 4)
            cell.value = f"=IF({prior_letter}{rev_r}=0,\"-\",({col_letter}{rev_r}-{prior_letter}{rev_r})/{prior_letter}{rev_r})"
            cell.font = F_FORMULA
            cell.number_format = PCT_FMT

    ws.cell(
        row=yoy_row + 2, column=1,
        value=(
            f"Source: SEC EDGAR XBRL companyfacts, {edgar.COMPANYFACTS_URL.format(cik10=cik10)} "
            "-- fallback-tag series via edgar.named_metric(); each quarter is the most-recently-FILED "
            "duration-matched (75-100 day) fact for that period end."
        ),
    ).font = F_SOURCE

    for i in range(len(ends_sorted) + 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = 13

    return {
        "ends": ends_sorted,
        "row_index": row_index,
        "rev_row": rev_r,
        "ni_row": ni_r,
        "eps_row": row_index["Diluted EPS"],
        "last_col": 1 + len(ends_sorted),
        "by_metric_by_end": by_metric_by_end,
    }


# --------------------------------------------------------------------------- #
# KPIs tab
# --------------------------------------------------------------------------- #


def build_kpis(wb: Workbook, ticker: str, entry: dict, is_ctx: dict, overrides: list[dict]) -> None:
    ws = _sheet(wb, "KPIs")
    ws["A1"] = f"{ticker} -- Company-specific KPIs (press-release-sourced)"
    ws["A1"].font = F_TITLE
    ws["A2"] = (
        "companyfacts (XBRL) does not carry segment/product/KPI detail -- only headline, "
        "non-dimensional financials (Session 2 finding, PLAN §3). These rows are TBD "
        "until extracted from the 8-K EX-99.1 press release and supplied via --kpi-overrides."
    )
    ws["A2"].font = F_SOURCE
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 60

    header_row = 4
    _header_row(ws, header_row, ["KPI", "Period", "Value", "Source"], start_col=1)
    kpis = [k for k in entry.get("key_kpi_concepts", []) if "(KPI)" in k or "KPI" in k]
    by_kpi: dict[str, list] = {}
    for o in overrides:
        if "metric" not in o:
            raise ValueError(f"--kpi-overrides entry missing required 'metric' key: {o!r}")
        by_kpi.setdefault(o["metric"], []).append(o)
    consumed: set[str] = set()

    r = header_row + 1
    if not kpis:
        ws.cell(row=r, column=1, value="(no KPI concepts listed in config/universe.json for this ticker)").font = F_SOURCE
        r += 1
        if overrides:
            print(
                f"  [build_model] warning: {ticker} has no key_kpi_concepts in config/universe.json, "
                f"but {len(overrides)} --kpi-overrides entries were supplied -- rendering them directly "
                "under their own 'metric' label since there is no universe.json KPI list to match against",
                file=sys.stderr,
            )
        for metric, matches in by_kpi.items():
            for m in matches:
                _kpi_row(ws, r, metric, m)
                r += 1
        return
    for kpi in kpis:
        matches = by_kpi.get(kpi, [{}])
        if kpi in by_kpi:
            consumed.add(kpi)
        for m in matches:
            _kpi_row(ws, r, kpi, m)
            r += 1

    unconsumed = set(by_kpi) - consumed
    if unconsumed:
        print(
            f"  [build_model] warning: {ticker} --kpi-overrides had {len(unconsumed)} entries whose "
            f"'metric' did not match any config/universe.json key_kpi_concepts entry: {sorted(unconsumed)} "
            "-- these values were supplied but NOT rendered anywhere in the workbook (fix the metric "
            "string to match universe.json exactly, or the KPI tab silently drops cited data)",
            file=sys.stderr,
        )


def _kpi_row(ws: Worksheet, row: int, label: str, m: dict) -> None:
    ws.cell(row=row, column=1, value=label).font = F_LABEL
    ws.cell(row=row, column=2, value=m.get("period", "")).font = F_INPUT
    val_cell = ws.cell(row=row, column=3)
    if "value" in m:
        val_cell.value = m["value"]
        val_cell.font = F_INPUT
        val_cell.number_format = CUR_FMT if m.get("unit") != "pct" else PCT_FMT
    else:
        val_cell.value = "TBD"
        val_cell.fill = YELLOW_FILL
    source = m.get("source_url", "")
    ws.cell(row=row, column=4, value=(f"{m.get('description', '')}: {source}" if source else "not yet extracted")).font = F_SOURCE


# --------------------------------------------------------------------------- #
# Our Estimates tab
# --------------------------------------------------------------------------- #


def build_estimates(wb: Workbook, ticker: str, is_ctx: dict) -> None:
    ws = _sheet(wb, "Our Estimates")
    ws["A1"] = f"{ticker} -- Our Estimate, Next Reported Period"
    ws["A1"].font = F_TITLE
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16

    last_col_letter = get_column_letter(is_ctx["last_col"])
    rev_row = is_ctx["rev_row"]

    r = 3
    ws.cell(row=r, column=1, value="Assumption").font = F_LABEL
    ws.cell(row=r, column=2, value="Value").font = F_HEADER
    ws.cell(row=r, column=2).fill = HEADER_FILL
    r += 1
    ws.cell(row=r, column=1, value="Assumed QoQ revenue growth").font = F_FORMULA
    growth_cell_row = r
    ws.cell(row=r, column=2, value=0.05).font = F_INPUT
    ws.cell(row=r, column=2).number_format = PCT_FMT
    r += 1
    ws.cell(row=r, column=1, value="Assumed net margin").font = F_FORMULA
    margin_cell_row = r
    ws.cell(row=r, column=2, value=0.10).font = F_INPUT
    ws.cell(row=r, column=2).number_format = PCT_FMT
    r += 2

    ws.cell(row=r, column=1, value="Last actual revenue ($M)").font = F_FORMULA
    last_rev_row = r
    c = ws.cell(row=r, column=2)
    c.value = f"='Income Statement'!{last_col_letter}{rev_row}"
    c.font = F_XREF
    c.number_format = CUR_FMT
    r += 1

    ws.cell(row=r, column=1, value="Estimated next-quarter revenue ($M)").font = F_LABEL
    est_rev_row = r
    c = ws.cell(row=r, column=2)
    c.value = f"=B{last_rev_row}*(1+B{growth_cell_row})"
    c.font = F_FORMULA
    c.number_format = CUR_FMT
    r += 1

    ws.cell(row=r, column=1, value="Estimated next-quarter net income ($M)").font = F_LABEL
    c = ws.cell(row=r, column=2)
    c.value = f"=B{est_rev_row}*B{margin_cell_row}"
    c.font = F_FORMULA
    c.number_format = CUR_FMT
    r += 2

    ws.cell(
        row=r, column=1,
        value=(
            "Assumption cells (blue) are ours to edit per-name before a /preview run; "
            "the estimate rows are live formulas, not hardcoded -- change an assumption "
            "and the estimate recalculates. This is a starting scaffold, not a forecast model."
        ),
    ).font = F_SOURCE


# --------------------------------------------------------------------------- #
# Pure-Python parallel check (no LibreOffice on this machine -- see module docstring)
# --------------------------------------------------------------------------- #


def _selfcheck_income_statement(facts_json: dict, ctx: dict) -> None:
    """Assert every value already placed on the Income Statement tab (ctx["by_metric_by_end"],
    anchored to Revenue's period-end set in build_income_statement) is finite, and that the
    margin/growth formulas built off them can never divide by zero. Validates the SAME data
    the sheet displays -- does not re-derive from EDGAR again (that would re-run the anchor
    logic and gain nothing beyond what build_income_statement already checked)."""
    rev_by_end = ctx["by_metric_by_end"][METRIC_ROWS[0][0]]
    for end in ctx["ends"]:
        rev = rev_by_end.get(end)
        if rev is not None and rev.val == 0:
            print(
                f"  [build_model] warning: revenue is exactly 0 at {end} -- margin formulas "
                "for this column will show '-' (guarded), not #DIV/0!",
                file=sys.stderr,
            )
    for label, by_end in ctx["by_metric_by_end"].items():
        for end, f in by_end.items():
            if not (f.val == f.val):  # NaN check
                raise ValueError(f"non-finite value in {label} at {end}")


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #


def build_model(
    ticker: str,
    *,
    n_periods: int = 8,
    kpi_overrides: list[dict] | None = None,
    out_path: Path | None = None,
) -> Path:
    ticker = ticker.upper()
    cik_info = edgar.resolve_cik(ticker)
    entry = load_universe_entry(ticker)
    facts_json = edgar.companyfacts(ticker)
    generated = datetime.now().strftime("%Y-%m-%d")

    wb = Workbook()
    wb.remove(wb.active)

    build_cover(wb, ticker, cik_info, entry, generated)
    is_ctx = build_income_statement(
        wb, facts_json, ticker, n_periods, entry.get("fiscal_year_end"), cik_info.cik10
    )
    _selfcheck_income_statement(facts_json, is_ctx)
    build_kpis(wb, ticker, entry, is_ctx, kpi_overrides or [])
    build_estimates(wb, ticker, is_ctx)

    out_path = out_path or (MODELS_DIR / ticker / f"{ticker}_model_{generated}.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(description="Build a Furton Coverage Excel model from EDGAR data")
    p.add_argument("ticker")
    p.add_argument("--periods", type=int, default=8, help="trailing quarters to show (default 8)")
    p.add_argument("--kpi-overrides", type=Path, default=None, help="JSON file of extracted KPI values")
    p.add_argument("--out", type=Path, default=None)
    args = p.parse_args(argv)

    overrides = []
    if args.kpi_overrides:
        overrides = json.loads(args.kpi_overrides.read_text(encoding="utf-8"))

    try:
        out = build_model(args.ticker, n_periods=args.periods, kpi_overrides=overrides, out_path=args.out)
    except edgar.EdgarError as exc:
        print(f"EDGAR ERROR: {exc}", file=sys.stderr)
        return 2
    print(f"wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(_main(sys.argv[1:]))
